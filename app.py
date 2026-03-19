import os
import csv
import io
import uuid
import shutil
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import database as db

# ── Initialise ──────────────────────────────────────────────────────

app = FastAPI(title="Device Inventory Tracker", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IS_VERCEL = os.environ.get("VERCEL", False)
UPLOAD_DIR = "/tmp/uploads" if IS_VERCEL else os.path.join(BASE_DIR, "uploads")
PHOTO_DIR = os.path.join(UPLOAD_DIR, "photos")
os.makedirs(PHOTO_DIR, exist_ok=True)

db.init_db()

# Mount static files
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
if not IS_VERCEL:
    app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


# ── Pages ───────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse(os.path.join(BASE_DIR, "static", "index.html"))


@app.get("/partner", response_class=HTMLResponse)
async def partner_page():
    return FileResponse(os.path.join(BASE_DIR, "static", "partner.html"))


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page():
    return FileResponse(os.path.join(BASE_DIR, "static", "dashboard.html"))


# ── CSV Upload (auto-assigns devices to partners) ──────────────────

@app.post("/api/upload-csv")
async def upload_csv(file: UploadFile = File(...)):
    if not file.filename.endswith(".csv"):
        raise HTTPException(400, "Only CSV files are accepted.")

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    required = {"partner_code", "partner_name", "device_id"}
    if not required.issubset(set(reader.fieldnames or [])):
        raise HTTPException(
            400,
            f"CSV must contain columns: {', '.join(sorted(required))}. "
            f"Found: {', '.join(reader.fieldnames or [])}",
        )

    partner_count = 0
    device_count = 0
    seen_partners = set()
    assignment_summary = {}  # partner_code -> device count

    for row in reader:
        pc = row["partner_code"].strip()
        pn = row["partner_name"].strip()
        city = row.get("city", "").strip()
        did = row["device_id"].strip()

        if not all([pc, pn, did]):
            continue

        if pc not in seen_partners:
            db.upsert_partner(pc, pn, city)
            seen_partners.add(pc)
            partner_count += 1
            assignment_summary[pc] = {"name": pn, "count": 0}

        db.upsert_device(pc, did)
        device_count += 1
        assignment_summary[pc]["count"] += 1

    # Save a copy
    save_path = os.path.join(UPLOAD_DIR, f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    with open(save_path, "wb") as f:
        f.write(content)

    return {
        "message": "CSV uploaded & devices auto-assigned to partners",
        "partners_imported": partner_count,
        "devices_imported": device_count,
        "assignments": assignment_summary,
    }


# ── Partner Login ───────────────────────────────────────────────────

@app.post("/api/partner/login")
async def partner_login(partner_code: str = Form(...)):
    partner = db.get_partner(partner_code.strip().upper())
    if not partner:
        raise HTTPException(404, "Partner code not found.")
    # Record login time
    db.record_partner_login(partner_code.strip().upper())
    # Return fresh partner data
    partner = db.get_partner(partner_code.strip().upper())
    devices = db.get_devices_by_partner(partner_code.strip().upper())
    return {
        **partner,
        "total_devices": len(devices),
    }


# ── Dummy Ping API ─────────────────────────────────────────────────

@app.post("/api/ping-check/{partner_code}")
async def ping_check(partner_code: str):
    partner = db.get_partner(partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")

    # Record that partner started verification
    db.record_verification_started(partner_code)

    db.simulate_ping(partner_code)
    devices = db.get_devices_by_partner(partner_code)

    online = [d for d in devices if d["is_online"]]
    offline = [d for d in devices if not d["is_online"]]

    return {
        "partner_code": partner_code,
        "total_devices": len(devices),
        "installed_count": len(online),
        "at_partner_count": len(offline),
        "installed_devices": [d["device_id"] for d in online],
        "at_partner_devices": [d["device_id"] for d in offline],
    }


# ── Get Partner Devices ────────────────────────────────────────────

@app.get("/api/partner/{partner_code}/devices")
async def partner_devices(partner_code: str):
    partner = db.get_partner(partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")
    devices = db.get_devices_by_partner(partner_code)
    return {"partner": partner, "devices": devices}


# ── Dummy OCR — Upload Photo & Extract Device IDs ──────────────────


def dummy_ocr(partner_code: str) -> list[str]:
    """Simulate OCR: returns 3-6 random offline device IDs for this partner."""
    import random
    offline = db.get_offline_devices(partner_code)
    unmatched = [d["device_id"] for d in offline if not d["ocr_matched"]]
    if not unmatched:
        return []
    count = min(random.randint(3, 6), len(unmatched))
    return random.sample(unmatched, count)


@app.post("/api/partner/upload-photo")
async def upload_photo_ocr(
    partner_code: str = Form(...),
    photo: UploadFile = File(...),
):
    partner = db.get_partner(partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")

    ext = os.path.splitext(photo.filename)[1] or ".jpg"
    photo_name = f"{partner_code}_ocr_{uuid.uuid4().hex[:8]}{ext}"
    photo_path = os.path.join(PHOTO_DIR, photo_name)
    with open(photo_path, "wb") as f:
        shutil.copyfileobj(photo.file, f)
    photo_url = f"/uploads/photos/{photo_name}"

    # Dummy OCR: simulate extracting device IDs from the photo
    extracted_ids = dummy_ocr(partner_code)
    matched_count = db.bulk_mark_ocr(partner_code, extracted_ids, photo_url)

    devices = db.get_devices_by_partner(partner_code)
    offline = [d for d in devices if not d["is_online"]]
    ocr_done = [d for d in offline if d["ocr_matched"]]
    remaining = [d for d in offline if not d["ocr_matched"]]

    return {
        "message": f"OCR detected {len(extracted_ids)} device(s) from your photo",
        "extracted_ids": extracted_ids,
        "matched_ids": extracted_ids,
        "unmatched_ocr_ids": [],
        "newly_matched": matched_count,
        "photo_url": photo_url,
        "total_at_partner": len(offline),
        "total_ocr_verified": len(ocr_done),
        "remaining_unverified": len(remaining),
        "remaining_ids": [d["device_id"] for d in remaining],
    }


# ── Submit Reason for Unverified Devices ───────────────────────────

class ReasonRequest(BaseModel):
    partner_code: str
    device_ids: list[str]
    reason: str


@app.post("/api/partner/submit-reason")
async def submit_reason(req: ReasonRequest):
    partner = db.get_partner(req.partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")

    if not req.reason.strip():
        raise HTTPException(400, "Reason cannot be empty.")

    updated = db.bulk_set_unverified_reason(req.partner_code, req.device_ids, req.reason.strip())

    return {
        "message": f"Reason submitted for {updated} device(s)",
        "updated_count": updated,
    }


# ── Ops Dashboard ───────────────────────────────────────────────────

@app.get("/api/ops/dashboard")
async def ops_dashboard():
    stats = db.get_dashboard_stats()
    counts = db.get_global_counts()
    return {"partners": stats, "counts": counts}


@app.get("/api/ops/partner/{partner_code}")
async def ops_partner_detail(partner_code: str):
    partner = db.get_partner(partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")
    devices = db.get_devices_by_partner(partner_code)
    return {"partner": partner, "devices": devices}


# ── Excel Export ────────────────────────────────────────────────────

@app.get("/api/ops/export/{partner_code}")
async def export_partner_report(partner_code: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    partner = db.get_partner(partner_code)
    if not partner:
        raise HTTPException(404, "Partner not found")

    devices = db.get_devices_by_partner(partner_code)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{partner_code} Inventory"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Device Inventory Report — {partner['partner_name']} ({partner_code})"
    ws["A1"].font = Font(bold=True, size=14, color="16213e")
    ws.append([])
    ws.append(["City:", partner.get("city", "N/A"), "", "Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])

    headers = ["Device ID", "Status", "OCR Verified", "OCR Photo", "OCR Matched At", "Reason (if unverified)"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_fills = {
        "installed": PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid"),
        "ocr_verified": PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid"),
        "unaccounted": PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid"),
    }

    for d in devices:
        if d["is_online"]:
            status = "Installed (Online)"
            fill_key = "installed"
        elif d["ocr_matched"]:
            status = "At Partner (OCR Verified)"
            fill_key = "ocr_verified"
        else:
            status = "At Partner (Unaccounted)"
            fill_key = "unaccounted"

        row_data = [
            d["device_id"],
            status,
            "Yes" if d["ocr_matched"] else "No",
            d.get("ocr_photo_path") or "—",
            d.get("ocr_matched_at") or "—",
            d.get("unverified_reason") or "—",
        ]
        ws.append(row_data)
        row_num = ws.max_row
        ws.cell(row=row_num, column=2).fill = status_fills.get(fill_key, PatternFill())
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        for col_num in range(1, 7):
            ws.cell(row=row_num, column=col_num).border = thin_border

    ws.append([])
    installed = sum(1 for d in devices if d["is_online"])
    ocr_verified = sum(1 for d in devices if not d["is_online"] and d["ocr_matched"])
    unaccounted = sum(1 for d in devices if not d["is_online"] and not d["ocr_matched"])
    ws.append(["Summary:", f"Total: {len(devices)}", f"Installed: {installed}", f"OCR Verified: {ocr_verified}", f"Unaccounted: {unaccounted}"])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"device_report_{partner_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/ops/export-all")
async def export_all_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    partners = db.get_all_partners()
    if not partners:
        raise HTTPException(404, "No data to export.")

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Partner Code", "Partner Name", "City", "Total", "Installed", "At Partner", "OCR Verified", "Unaccounted"])
    for col_num in range(1, 9):
        cell = summary_ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for p in partners:
        devices = db.get_devices_by_partner(p["partner_code"])
        installed = sum(1 for d in devices if d["is_online"])
        at_partner = sum(1 for d in devices if not d["is_online"])
        ocr_verified = sum(1 for d in devices if not d["is_online"] and d["ocr_matched"])
        unaccounted = sum(1 for d in devices if not d["is_online"] and not d["ocr_matched"])
        summary_ws.append([p["partner_code"], p["partner_name"], p.get("city", ""), len(devices), installed, at_partner, ocr_verified, unaccounted])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"device_full_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Clear Data (dev utility) ───────────────────────────────────────

@app.post("/api/ops/clear-data")
async def clear_data():
    db.clear_all_data()
    return {"message": "All data cleared."}


# ── Run ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
